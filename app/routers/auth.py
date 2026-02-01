from typing import List

from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel

from ..database import get_db_client, ROLE_TABLES
from ..schemas import (
    LoginRequest,
    LoginResponse,
    UserCreate,
    UserProfileUpdate,
    UserResponse,
    UserUpdate,
)
from ..utils import get_password_hash, verify_password

router = APIRouter(tags=["Authentication & Users"])


def _esc(s: str) -> str:
    return (s or "").replace("'", "''")


def _name_from_parts(first_name: str | None, middle_name: str | None, last_name: str | None) -> str:
    parts = [p for p in (first_name, middle_name, last_name) if p and str(p).strip()]
    return " ".join(parts).strip() if parts else ""


@router.post("/login", response_model=LoginResponse)
def login(creds: LoginRequest):
    client = get_db_client()
    try:
        result = client.sqlQuery(
            f"SELECT hashed_password, role, active FROM users WHERE username = '{_esc(creds.username)}'"
        )
        if not result:
            return {"success": False, "message": "User not found"}

        stored_hash, role, is_active = result[0][0], result[0][1], result[0][2]
        if not is_active:
            return {"success": False, "message": "Account is disabled"}

        if not verify_password(creds.password, stored_hash):
            return {"success": False, "message": "Invalid password"}

        un = _esc(creds.username)
        name = ""
        first_name = last_name = strand = payment_plan = None

        if role == "student":
            row = client.sqlQuery(f"SELECT first_name, middle_name, last_name, gender, strand, payment_plan FROM students WHERE username = '{un}'")
            if row:
                r = row[0]
                first_name = r[0] if len(r) > 0 else None
                last_name = r[2] if len(r) > 2 else None
                name = _name_from_parts(r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                strand = r[4] if len(r) > 4 else None
                payment_plan = r[5] if len(r) > 5 else None
        elif role == "staff":
            row = client.sqlQuery(f"SELECT first_name, middle_name, last_name FROM staff WHERE username = '{un}'")
            if row:
                r = row[0]
                first_name = r[0] if len(r) > 0 else None
                last_name = r[2] if len(r) > 2 else None
                name = _name_from_parts(r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
        elif role in ROLE_TABLES:
            row = client.sqlQuery(f"SELECT first_name, middle_name, last_name, gender, contact_information FROM {role} WHERE username = '{un}'")
            if row:
                r = row[0]
                first_name = r[0] if len(r) > 0 else None
                last_name = r[2] if len(r) > 2 else None
                name = _name_from_parts(r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)

        if not name:
            name = creds.username

        return {
            "success": True,
            "role": role,
            "name": name,
            "first_name": first_name,
            "last_name": last_name,
            "strand": strand,
            "payment_plan": payment_plan,
            "message": "Login successful",
        }
    except Exception as e:
        print(f"Login Error: {e}")
        return {"success": False, "message": "Login failed"}


# --- USER MANAGEMENT ---


def _fetch_role_row(client, username: str, role: str):
    un = _esc(username)
    if role == "student":
        r = client.sqlQuery(f"SELECT first_name, middle_name, last_name, gender, strand, section, payment_plan FROM students WHERE username = '{un}'")
        if r:
            row = r[0]
            return {
                "first_name": row[0], "middle_name": row[1], "last_name": row[2],
                "gender": row[3] if len(row) > 3 else None,
                "strand": row[4] if len(row) > 4 else None,
                "section": row[5] if len(row) > 5 else None,
                "payment_plan": row[6] if len(row) > 6 else None,
                "contact_info": None, "gen_role": None,
                "position": None, "department": None, "date_hired": None, "status": None, "monthly_salary": None,
            }
    elif role == "staff":
        r = client.sqlQuery(f"SELECT first_name, middle_name, last_name, gender, position, department, date_hired, status, monthly_salary FROM staff WHERE username = '{un}'")
        if r:
            row = r[0]
            return {
                "first_name": row[0], "middle_name": row[1], "last_name": row[2],
                "gender": row[3] if len(row) > 3 else None,
                "position": row[4] if len(row) > 4 else None,
                "department": row[5] if len(row) > 5 else None,
                "date_hired": row[6] if len(row) > 6 else None,
                "status": row[7] if len(row) > 7 else None,
                "monthly_salary": row[8] if len(row) > 8 else None,
                "strand": None, "section": None, "payment_plan": None,
                "contact_info": None, "gen_role": None,
            }
    elif role in ROLE_TABLES:
        r = client.sqlQuery(f"SELECT first_name, middle_name, last_name, gender, contact_information FROM {role} WHERE username = '{un}'")
        if r:
            row = r[0]
            return {
                "first_name": row[0], "middle_name": row[1], "last_name": row[2],
                "gender": row[3] if len(row) > 3 else None,
                "contact_info": row[4] if len(row) > 4 else None,
                "strand": None, "section": None, "payment_plan": None,
                "position": None, "department": None, "date_hired": None, "status": None, "monthly_salary": None,
                "gen_role": None,
            }
    return None


@router.get("/users", response_model=List[UserResponse])
def get_users():
    """List all users (users + role table data)."""
    client = get_db_client()
    try:
        result = client.sqlQuery("SELECT username, role, active FROM users")
        users = []
        for row in result:
            username, role, active = row[0], row[1], row[2]
            profile = _fetch_role_row(client, username, role)
            if profile:
                name = _name_from_parts(profile.get("first_name"), profile.get("middle_name"), profile.get("last_name")) or username
            else:
                name = username
            user_data = {
                "username": username,
                "role": role,
                "name": name,
                "active": active,
                "first_name": profile.get("first_name") if profile else None,
                "middle_name": profile.get("middle_name") if profile else None,
                "last_name": profile.get("last_name") if profile else None,
                "contact_info": profile.get("contact_info") if profile else None,
                "gender": profile.get("gender") if profile else None,
                "strand": profile.get("strand") if profile else None,
                "section": profile.get("section") if profile else None,
                "payment_plan": profile.get("payment_plan") if profile else None,
                "gen_role": profile.get("gen_role") if profile else None,
                "position": profile.get("position") if profile else None,
                "department": profile.get("department") if profile else None,
                "date_hired": profile.get("date_hired") if profile else None,
                "status": profile.get("status") if profile else None,
                "monthly_salary": profile.get("monthly_salary") if profile else None,
            }
            users.append(user_data)
        return users
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/users", response_model=UserResponse)
def create_user(user: UserCreate):
    """Create a new user (users + role table row)."""
    client = get_db_client()
    try:
        check = client.sqlQuery(f"SELECT id FROM users WHERE username = '{_esc(user.username)}'")
        if check:
            raise HTTPException(status_code=400, detail="Username already exists")

        password_to_use = "123" if user.role == "student" else user.password
        hashed_pw = get_password_hash(password_to_use)

        client.sqlExec(
            f"INSERT INTO users (username, hashed_password, role, active) "
            f"VALUES ('{_esc(user.username)}', '{_esc(hashed_pw)}', '{_esc(user.role)}', {str(user.active).lower()})"
        )

        fn, mn, ln = user.first_name or "", user.middle_name or "", user.last_name or ""
        if user.role == "student":
            section = getattr(user, "section", None) or ""
            payment_plan = (user.payment_plan or "plan_a") if getattr(user, "payment_plan", None) else "plan_a"
            client.sqlExec(
                f"INSERT INTO students (username, first_name, middle_name, last_name, gender, strand, section, payment_plan) "
                f"VALUES ('{_esc(user.username)}', '{_esc(fn)}', '{_esc(mn)}', '{_esc(ln)}', "
                f"'{_esc(user.gender or "")}', '{_esc(user.strand or "")}', '{_esc(section)}', '{_esc(payment_plan)}')"
            )
        elif user.role == "staff":
            pos = getattr(user, "position", None) or ""
            dept = getattr(user, "department", None) or ""
            dh = getattr(user, "date_hired", None) or ""
            st = getattr(user, "status", None) or ""
            sal = getattr(user, "monthly_salary", None)
            sal_val = int(sal) if sal is not None else 0
            client.sqlExec(
                f"INSERT INTO staff (username, first_name, middle_name, last_name, gender, position, department, date_hired, status, monthly_salary) "
                f"VALUES ('{_esc(user.username)}', '{_esc(fn)}', '{_esc(mn)}', '{_esc(ln)}', '{_esc(user.gender or "")}', "
                f"'{_esc(pos)}', '{_esc(dept)}', '{_esc(dh)}', '{_esc(st)}', {sal_val})"
            )
        elif user.role in ROLE_TABLES:
            contact = user.contact_info or ""
            client.sqlExec(
                f"INSERT INTO {user.role} (username, first_name, middle_name, last_name, gender, contact_information) "
                f"VALUES ('{_esc(user.username)}', '{_esc(fn)}', '{_esc(mn)}', '{_esc(ln)}', "
                f"'{_esc(user.gender or "")}', '{_esc(contact)}')"
            )

        name = _name_from_parts(user.first_name, user.middle_name, user.last_name) or user.username
        resp = {
            "username": user.username,
            "role": user.role,
            "name": name,
            "active": user.active,
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "contact_info": user.contact_info,
            "gender": user.gender,
            "strand": user.strand,
            "section": getattr(user, "section", None),
            "payment_plan": user.payment_plan,
            "gen_role": None,
            "position": getattr(user, "position", None),
            "department": getattr(user, "department", None),
            "date_hired": getattr(user, "date_hired", None),
            "status": getattr(user, "status", None),
            "monthly_salary": getattr(user, "monthly_salary", None),
        }
        return resp
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/users/{username}")
def update_user(username: str, user: UserUpdate):
    """Update user (users + role table)."""
    client = get_db_client()
    try:
        check = client.sqlQuery(f"SELECT id, role FROM users WHERE username = '{_esc(username)}'")
        if not check:
            raise HTTPException(status_code=404, detail="User not found")
        current_role = check[0][1]
        un = _esc(username)

        updates = []
        if user.role:
            updates.append(f"role = '{_esc(user.role)}'")
        if user.active is not None:
            updates.append(f"active = {str(user.active).lower()}")
        if user.password:
            updates.append(f"hashed_password = '{_esc(get_password_hash(user.password))}'")
        if updates:
            client.sqlExec(f"UPDATE users SET {', '.join(updates)} WHERE username = '{un}'")

        # Update role table (current role)
        role = user.role or current_role
        if role == "student":
            up = []
            if user.first_name is not None:
                up.append(f"first_name = '{_esc(user.first_name)}'")
            if user.middle_name is not None:
                up.append(f"middle_name = '{_esc(user.middle_name)}'")
            if user.last_name is not None:
                up.append(f"last_name = '{_esc(user.last_name)}'")
            if user.gender is not None:
                up.append(f"gender = '{_esc(user.gender)}'")
            if user.strand is not None:
                up.append(f"strand = '{_esc(user.strand)}'")
            if getattr(user, "section", None) is not None:
                up.append(f"section = '{_esc(user.section)}'")
            if user.payment_plan is not None:
                if user.payment_plan not in ("", "plan_a", "plan_b", "plan_c"):
                    raise HTTPException(status_code=400, detail="payment_plan must be plan_a, plan_b, or plan_c")
                up.append(f"payment_plan = '{_esc(user.payment_plan)}'")
            if up:
                client.sqlExec(f"UPDATE students SET {', '.join(up)} WHERE username = '{un}'")
        elif role == "staff":
            up = []
            if user.first_name is not None:
                up.append(f"first_name = '{_esc(user.first_name)}'")
            if user.middle_name is not None:
                up.append(f"middle_name = '{_esc(user.middle_name)}'")
            if user.last_name is not None:
                up.append(f"last_name = '{_esc(user.last_name)}'")
            if user.gender is not None:
                up.append(f"gender = '{_esc(user.gender)}'")
            if getattr(user, "position", None) is not None:
                up.append(f"position = '{_esc(user.position)}'")
            if getattr(user, "department", None) is not None:
                up.append(f"department = '{_esc(user.department)}'")
            if getattr(user, "date_hired", None) is not None:
                up.append(f"date_hired = '{_esc(user.date_hired)}'")
            if getattr(user, "status", None) is not None:
                up.append(f"status = '{_esc(user.status)}'")
            if getattr(user, "monthly_salary", None) is not None:
                up.append(f"monthly_salary = {int(user.monthly_salary)}")
            if up:
                client.sqlExec(f"UPDATE staff SET {', '.join(up)} WHERE username = '{un}'")
            # Staff deductions (payroll): replace all for this staff
            deductions = getattr(user, "deductions", None)
            if deductions is not None and isinstance(deductions, list) and len(deductions) >= 0:
                try:
                    client.sqlExec(f"DELETE FROM staff_deductions WHERE staff_id = '{un}'")
                except Exception:
                    pass
                for d in deductions:
                    dtype = (getattr(d, "deduction_type", None) or (d.get("deduction_type") if isinstance(d, dict) else None) or "").strip() or "Deduction"
                    amt = getattr(d, "amount", None)
                    if amt is None and isinstance(d, dict):
                        amt = d.get("amount", 0)
                    amt = float(amt or 0)
                    amt_cents = int(round(amt * 100))
                    client.sqlExec(
                        f"INSERT INTO staff_deductions (staff_id, deduction_type, amount) VALUES ('{un}', '{_esc(str(dtype))}', {amt_cents})"
                    )
        elif role in ROLE_TABLES:
            up = []
            if user.first_name is not None:
                up.append(f"first_name = '{_esc(user.first_name)}'")
            if user.middle_name is not None:
                up.append(f"middle_name = '{_esc(user.middle_name)}'")
            if user.last_name is not None:
                up.append(f"last_name = '{_esc(user.last_name)}'")
            if user.gender is not None:
                up.append(f"gender = '{_esc(user.gender)}'")
            if user.contact_info is not None:
                up.append(f"contact_information = '{_esc(user.contact_info)}'")
            if up:
                client.sqlExec(f"UPDATE {role} SET {', '.join(up)} WHERE username = '{un}'")

        return {"message": "User updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ProfileRequest(BaseModel):
    username: str


def _get_user_profile(username: str):
    """Fetch user profile from users + role table."""
    client = get_db_client()
    if not username:
        raise HTTPException(status_code=400, detail="Username is required")
    un = _esc(username)
    result = client.sqlQuery(f"SELECT username, role FROM users WHERE username = '{un}'")
    if not result:
        raise HTTPException(status_code=404, detail=f"User '{username}' not found")
    role = result[0][1]
    profile = _fetch_role_row(client, username, role)
    if not profile:
        return {"username": username, "role": role, "name": username, "first_name": None, "middle_name": None, "last_name": None, "contact_info": None, "gender": None, "strand": None, "payment_plan": None, "gen_role": None, "position": None, "department": None, "date_hired": None, "status": None, "monthly_salary": None}
    name = _name_from_parts(profile.get("first_name"), profile.get("middle_name"), profile.get("last_name")) or username
    out = {
        "username": username,
        "role": role,
        "name": name,
        "first_name": profile.get("first_name"),
        "middle_name": profile.get("middle_name"),
        "last_name": profile.get("last_name"),
        "contact_info": profile.get("contact_info"),
        "gender": profile.get("gender"),
        "strand": profile.get("strand"),
        "payment_plan": profile.get("payment_plan"),
        "gen_role": profile.get("gen_role"),
    }
    if role == "staff":
        out["position"] = profile.get("position")
        out["department"] = profile.get("department")
        out["date_hired"] = profile.get("date_hired")
        out["status"] = profile.get("status")
        out["monthly_salary"] = profile.get("monthly_salary")
    return out


@router.post("/profile")
def get_current_user_profile(req: ProfileRequest):
    """Get current user's profile."""
    try:
        return _get_user_profile(req.username)
    except HTTPException:
        raise
    except Exception as e:
        print(f"Profile fetch error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ProfileUpdateRequest(BaseModel):
    username: str
    profile: UserProfileUpdate


@router.put("/profile")
def update_user_profile(req: ProfileUpdateRequest):
    """Update current user's own profile (role table only)."""
    client = get_db_client()
    username = req.username
    profile = req.profile
    try:
        result = client.sqlQuery(f"SELECT role FROM users WHERE username = '{_esc(username)}'")
        if not result:
            raise HTTPException(status_code=404, detail="User not found")
        role = result[0][0]
        un = _esc(username)

        if role == "student":
            up = []
            if profile.first_name is not None:
                up.append(f"first_name = '{_esc(profile.first_name)}'")
            if profile.middle_name is not None:
                up.append(f"middle_name = '{_esc(profile.middle_name)}'")
            if profile.last_name is not None:
                up.append(f"last_name = '{_esc(profile.last_name)}'")
            if profile.contact_info is not None:
                pass  # students don't have contact_info in table; we use section or leave it
            if profile.gender is not None:
                up.append(f"gender = '{_esc(profile.gender)}'")
            if profile.payment_plan is not None:
                if profile.payment_plan not in ("", "plan_a", "plan_b", "plan_c"):
                    raise HTTPException(status_code=400, detail="payment_plan must be plan_a, plan_b, or plan_c")
                up.append(f"payment_plan = '{_esc(profile.payment_plan)}'")
            if profile.strand is not None:
                up.append(f"strand = '{_esc(profile.strand)}'")
            if getattr(profile, "section", None) is not None:
                up.append(f"section = '{_esc(profile.section)}'")
            if up:
                client.sqlExec(f"UPDATE students SET {', '.join(up)} WHERE username = '{un}'")
        elif role == "staff":
            up = []
            if profile.first_name is not None:
                up.append(f"first_name = '{_esc(profile.first_name)}'")
            if profile.middle_name is not None:
                up.append(f"middle_name = '{_esc(profile.middle_name)}'")
            if profile.last_name is not None:
                up.append(f"last_name = '{_esc(profile.last_name)}'")
            if profile.gender is not None:
                up.append(f"gender = '{_esc(profile.gender)}'")
            if up:
                client.sqlExec(f"UPDATE staff SET {', '.join(up)} WHERE username = '{un}'")
        elif role in ROLE_TABLES:
            up = []
            if profile.first_name is not None:
                up.append(f"first_name = '{_esc(profile.first_name)}'")
            if profile.middle_name is not None:
                up.append(f"middle_name = '{_esc(profile.middle_name)}'")
            if profile.last_name is not None:
                up.append(f"last_name = '{_esc(profile.last_name)}'")
            if profile.gender is not None:
                up.append(f"gender = '{_esc(profile.gender)}'")
            if profile.contact_info is not None:
                up.append(f"contact_information = '{_esc(profile.contact_info)}'")
            if up:
                client.sqlExec(f"UPDATE {role} SET {', '.join(up)} WHERE username = '{un}'")

        return {"message": "Profile updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class PasswordChangeRequest(BaseModel):
    username: str
    current_password: str
    new_password: str


@router.put("/change-password")
def change_password(req: PasswordChangeRequest):
    """Change current user's password."""
    client = get_db_client()
    try:
        result = client.sqlQuery(f"SELECT hashed_password FROM users WHERE username = '{_esc(req.username)}'")
        if not result:
            raise HTTPException(status_code=404, detail="User not found")
        if not verify_password(req.current_password, result[0][0]):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        new_hash = get_password_hash(req.new_password)
        client.sqlExec(f"UPDATE users SET hashed_password = '{_esc(new_hash)}' WHERE username = '{_esc(req.username)}'")
        return {"message": "Password changed successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _normalize_header(h: str) -> str:
    if not h:
        return ""
    return str(h).strip().lower().replace(" ", "_").replace("-", "_")


@router.post("/users/import-students")
async def import_students_excel(file: UploadFile = File(...)):
    """
    Import students from Excel. Creates users (role=student) + students rows.
    Expected columns: student_id/username, last_name, first_name, middle_name, gender, strand, section, contact_information, payment_plan.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx)")

    try:
        import io
        from openpyxl import load_workbook

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="Excel file has no sheet")
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel must have a header row and at least one data row")

        headers = [_normalize_header(str(c)) if c is not None else "" for c in rows[0]]
        col_map = {}
        for name in ("student_id", "username", "last_name", "first_name", "middle_name", "gender", "strand", "section", "contact_information", "contact_info", "payment_plan"):
            for i, h in enumerate(headers):
                if h == name or h == name.replace("_", ""):
                    col_map[name] = i
                    break
        if "contact_information" not in col_map and "contact_info" in col_map:
            col_map["contact_information"] = col_map["contact_info"]
        if "student_id" not in col_map and "username" in col_map:
            col_map["student_id"] = col_map["username"]
        if "student_id" not in col_map:
            raise HTTPException(status_code=400, detail="Excel must have a 'student_id' or 'username' column")

        def get_cell(row, key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        client = get_db_client()
        hashed_pw = get_password_hash("123")
        created = 0
        skipped = []
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(v is not None and str(v).strip() for v in row):
                continue
            username = get_cell(row, "student_id") or get_cell(row, "username")
            if not username:
                errors.append(f"Row {row_idx}: missing student_id/username")
                continue
            last_name = get_cell(row, "last_name")
            first_name = get_cell(row, "first_name")
            middle_name = get_cell(row, "middle_name")
            gender = get_cell(row, "gender")
            strand_val = get_cell(row, "strand")
            section_val = get_cell(row, "section")
            payment_plan_val = get_cell(row, "payment_plan") or "plan_a"
            if payment_plan_val not in ("plan_a", "plan_b", "plan_c"):
                payment_plan_val = "plan_a"
            try:
                check = client.sqlQuery(f"SELECT id FROM users WHERE username = '{_esc(username)}'")
                if check:
                    skipped.append(username)
                    continue
                client.sqlExec(
                    f"INSERT INTO users (username, hashed_password, role, active) "
                    f"VALUES ('{_esc(username)}', '{hashed_pw}', 'student', true)"
                )
                client.sqlExec(
                    f"INSERT INTO students (username, first_name, middle_name, last_name, gender, strand, section, payment_plan) "
                    f"VALUES ('{_esc(username)}', '{_esc(first_name)}', '{_esc(middle_name)}', '{_esc(last_name)}', "
                    f"'{_esc(gender)}', '{_esc(strand_val)}', '{_esc(section_val)}', '{_esc(payment_plan_val)}')"
                )
                created += 1
            except Exception as e:
                errors.append(f"Row {row_idx} ({username}): {str(e)}")

        return {
            "success": True,
            "created": created,
            "skipped_usernames": skipped,
            "errors": errors,
            "message": f"Created {created} student(s). Skipped {len(skipped)} existing. {len(errors)} error(s).",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


@router.delete("/users/{username}")
def delete_user(username: str):
    """Delete a user (remove from role table then users)."""
    client = get_db_client()
    try:
        check = client.sqlQuery(f"SELECT id, role FROM users WHERE username = '{_esc(username)}'")
        if not check:
            raise HTTPException(status_code=404, detail="User not found")
        role = check[0][1]
        un = _esc(username)
        if role == "student":
            client.sqlExec(f"DELETE FROM students WHERE username = '{un}'")
        elif role == "staff":
            client.sqlExec(f"DELETE FROM staff WHERE username = '{un}'")
        elif role in ROLE_TABLES:
            client.sqlExec(f"DELETE FROM {role} WHERE username = '{un}'")
        client.sqlExec(f"DELETE FROM users WHERE username = '{un}'")
        return {"message": "User deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
